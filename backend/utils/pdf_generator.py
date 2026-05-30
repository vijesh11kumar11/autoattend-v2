"""
AutoAttend AI v2.0 — PDF / Excel generators

Functions:
  generate_student_pdf(student_id, date_from, date_to, db) → BytesIO
  generate_class_session_pdf(session_id, db)               → BytesIO
  generate_defaulters_pdf(course_ids, semester, threshold, db) → BytesIO
  generate_monthly_excel(subject_id, year, month, db)      → BytesIO
"""

import calendar
import io
from datetime import date, datetime
from typing import Optional

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    HRFlowable,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from sqlalchemy import case, func
from sqlalchemy.orm import Session

# ── Palette ───────────────────────────────────────────────────────────
C_PRIMARY = colors.HexColor("#1e40af")  # dark-blue
C_HEADER = colors.HexColor("#1e3a5f")
C_ROW_ALT = colors.HexColor("#f0f4ff")
C_GREEN = colors.HexColor("#16a34a")
C_AMBER = colors.HexColor("#d97706")
C_RED = colors.HexColor("#dc2626")
C_GREY = colors.HexColor("#94a3b8")
C_WHITE = colors.white
C_BLACK = colors.HexColor("#0f172a")

STYLES = getSampleStyleSheet()
TITLE_STYLE = ParagraphStyle(
    "title",
    parent=STYLES["Title"],
    fontSize=16,
    textColor=C_HEADER,
    spaceAfter=4,
)
SUBTITLE_STYLE = ParagraphStyle(
    "subtitle",
    parent=STYLES["Normal"],
    fontSize=9,
    textColor=C_GREY,
    spaceAfter=12,
)
BODY_STYLE = ParagraphStyle(
    "body",
    parent=STYLES["Normal"],
    fontSize=9,
    textColor=C_BLACK,
)


def _status_color(pct: float, threshold: float) -> colors.Color:
    if pct >= threshold:
        return C_GREEN
    if pct >= threshold - 10:
        return C_AMBER
    return C_RED


def _pct_str(present: int, total: int) -> str:
    if not total:
        return "—"
    return f"{round(present * 100 / total, 1)}%"


def _base_doc(buf: io.BytesIO, title: str, landscape_mode: bool = False) -> SimpleDocTemplate:
    ps = landscape(A4) if landscape_mode else A4
    return SimpleDocTemplate(
        buf,
        pagesize=ps,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
        title=title,
    )


def _table_style(has_total_row: bool = False) -> TableStyle:
    cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), C_PRIMARY),
        ("TEXTCOLOR", (0, 0), (-1, 0), C_WHITE),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [C_WHITE, C_ROW_ALT]),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.3, C_GREY),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    return TableStyle(cmds)


def _header_block(story: list, title: str, meta_lines: list[str]):
    story.append(Paragraph(title, TITLE_STYLE))
    for line in meta_lines:
        story.append(Paragraph(line, SUBTITLE_STYLE))
    story.append(HRFlowable(width="100%", thickness=1, color=C_PRIMARY, spaceAfter=10))


# ═══════════════════════════════════════════════════════════════════════
# 1. Student Attendance PDF
# ═══════════════════════════════════════════════════════════════════════


def generate_student_pdf(
    student_id: int,
    date_from: Optional[date],
    date_to: Optional[date],
    db: Session,
) -> io.BytesIO:
    from config import settings
    from database import (
        AttendanceRecord,
        AttendanceSession,
        AttendanceStatus,
        Course,
        SessionStatus,
        Subject,
        User,
        UserRole,
    )

    student = db.query(User).filter(User.id == student_id, User.role == UserRole.student).first()
    if not student:
        raise ValueError(f"Student {student_id} not found.")

    course = db.query(Course).filter(Course.id == student.course_id).first()

    # Ended sessions for filtering
    session_q = db.query(AttendanceSession).filter(AttendanceSession.status == SessionStatus.ended)
    if date_from:
        session_q = session_q.filter(AttendanceSession.date >= date_from)
    if date_to:
        session_q = session_q.filter(AttendanceSession.date <= date_to)
    sessions = {s.id: s for s in session_q.all()}

    # Per-subject stats
    rows = (
        db.query(
            Subject.name.label("subject_name"),
            Subject.code,
            func.count(AttendanceRecord.id).label("total"),
            func.sum(case((AttendanceRecord.status == AttendanceStatus.present, 1), else_=0)).label(
                "present"
            ),
            func.sum(case((AttendanceRecord.status == AttendanceStatus.absent, 1), else_=0)).label(
                "absent"
            ),
        )
        .join(AttendanceSession, AttendanceRecord.session_id == AttendanceSession.id)
        .join(Subject, AttendanceSession.subject_id == Subject.id)
        .filter(
            AttendanceRecord.student_id == student_id,
            AttendanceRecord.session_id.in_(list(sessions.keys())),
        )
        .group_by(Subject.id, Subject.name, Subject.code)
        .order_by(Subject.name)
        .all()
    )

    buf = io.BytesIO()
    doc = _base_doc(buf, f"Attendance Report – {student.name}")
    story = []

    date_range = ""
    if date_from or date_to:
        df = str(date_from) if date_from else "—"
        dt = str(date_to) if date_to else "—"
        date_range = f"Period: {df} to {dt}"

    _header_block(
        story,
        "Student Attendance Report",
        [
            f"Name: {student.name}   |   Roll No: {student.roll_number or '—'}",
            f"Course: {course.name if course else '—'}   |   Semester: {student.semester or '—'}",
            date_range or f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}",
        ],
    )

    # Summary table
    headers = ["Subject", "Code", "Present", "Absent", "Total", "Attendance %", "Status"]
    data = [headers]
    total_p = total_a = total_t = 0

    for r in rows:
        pct = round(r.present * 100 / r.total, 1) if r.total else 0.0
        total_p += r.present
        total_a += r.absent
        total_t += r.total
        st = (
            "SAFE"
            if pct >= settings.ATTENDANCE_THRESHOLD
            else (
                "WARNING"
                if pct >= settings.ATTENDANCE_THRESHOLD - 10
                else ("CRITICAL" if pct >= settings.ATTENDANCE_THRESHOLD - 25 else "DETAINED")
            )
        )
        data.append([r.subject_name, r.code, r.present, r.absent, r.total, f"{pct}%", st])

    overall_pct = round(total_p * 100 / total_t, 1) if total_t else 0.0
    data.append(["TOTAL", "", total_p, total_a, total_t, f"{overall_pct}%", ""])

    col_widths = [6 * cm, 2.5 * cm, 1.8 * cm, 1.8 * cm, 1.8 * cm, 2.8 * cm, 2.5 * cm]
    tbl = Table(data, colWidths=col_widths, repeatRows=1)
    ts = _table_style()
    # Color % column based on value
    for i, row in enumerate(data[1:-1], start=1):
        try:
            pct_val = float(str(row[5]).replace("%", ""))
        except ValueError:
            continue
        c = _status_color(pct_val, settings.ATTENDANCE_THRESHOLD)
        ts.add("TEXTCOLOR", (5, i), (5, i), c)
        ts.add("FONTNAME", (5, i), (5, i), "Helvetica-Bold")

    # Total row
    last = len(data) - 1
    ts.add("BACKGROUND", (0, last), (-1, last), C_HEADER)
    ts.add("TEXTCOLOR", (0, last), (-1, last), C_WHITE)
    ts.add("FONTNAME", (0, last), (-1, last), "Helvetica-Bold")
    tbl.setStyle(ts)
    story.append(tbl)

    story.append(Spacer(1, 0.5 * cm))
    story.append(
        Paragraph(
            f"Overall Attendance: <b>{overall_pct}%</b>   |   "
            f"Threshold: <b>{settings.ATTENDANCE_THRESHOLD}%</b>   |   "
            f"{'⚠ Below threshold' if overall_pct < settings.ATTENDANCE_THRESHOLD else '✓ Above threshold'}",
            BODY_STYLE,
        )
    )

    doc.build(story)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════════════════
# 2. Class Session PDF
# ═══════════════════════════════════════════════════════════════════════


def generate_class_session_pdf(session_id: int, db: Session) -> io.BytesIO:
    from database import (
        AttendanceRecord,
        AttendanceSession,
        Subject,
        User,
    )

    sess = db.query(AttendanceSession).filter(AttendanceSession.id == session_id).first()
    if not sess:
        raise ValueError(f"Session {session_id} not found.")

    subj = db.query(Subject).filter(Subject.id == sess.subject_id).first()
    teacher = db.query(User).filter(User.id == sess.teacher_id).first()

    records = (
        db.query(AttendanceRecord, User.name.label("student_name"), User.roll_number)
        .join(User, AttendanceRecord.student_id == User.id)
        .filter(AttendanceRecord.session_id == session_id)
        .order_by(User.roll_number)
        .all()
    )

    buf = io.BytesIO()
    doc = _base_doc(buf, f"Class Session Report – {subj.name if subj else session_id}")
    story = []

    _header_block(
        story,
        "Class Session Attendance Report",
        [
            f"Subject: {subj.name if subj else '—'} ({subj.code if subj else '—'})   |   "
            f"Teacher: {teacher.name if teacher else '—'}",
            f"Date: {sess.date}   |   "
            f"Time: {sess.start_time} – {sess.end_time}   |   "
            f"Present: {sess.present_count}/{sess.total_students}",
        ],
    )

    data = [["#", "Roll No", "Name", "Status", "Method", "Face Verified", "Marked At"]]
    for idx, (rec, name, roll) in enumerate(records, 1):
        data.append(
            [
                idx,
                roll or "—",
                name or "—",
                rec.status.value.title(),
                (rec.marked_by or "—").replace("_", " ").title(),
                "Yes" if rec.face_verified else "No",
                rec.marked_at.strftime("%H:%M:%S") if rec.marked_at else "—",
            ]
        )

    col_widths = [1 * cm, 2.5 * cm, 5.5 * cm, 2 * cm, 2.8 * cm, 2.5 * cm, 2 * cm]
    tbl = Table(data, colWidths=col_widths, repeatRows=1)
    ts = _table_style()
    # Colour status column
    for i, row in enumerate(data[1:], start=1):
        status_val = str(row[3]).lower()
        c = C_GREEN if "present" in status_val else (C_AMBER if "late" in status_val else C_RED)
        ts.add("TEXTCOLOR", (3, i), (3, i), c)
        ts.add("FONTNAME", (3, i), (3, i), "Helvetica-Bold")
    tbl.setStyle(ts)
    story.append(tbl)

    doc.build(story)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════════════════
# 3. Defaulters PDF
# ═══════════════════════════════════════════════════════════════════════


def generate_defaulters_pdf(
    course_ids: list[int],
    semester: Optional[int],
    threshold: float,
    db: Session,
) -> io.BytesIO:
    from database import (
        AttendanceRecord,
        AttendanceSession,
        AttendanceStatus,
        SessionStatus,
        Subject,
        User,
        UserRole,
    )

    # Get subjects filtered by semester
    subj_q = db.query(Subject).filter(Subject.course_id.in_(course_ids))
    if semester is not None:
        subj_q = subj_q.filter(Subject.semester == semester)
    subjects = subj_q.all()
    subject_ids = [s.id for s in subjects]

    student_ids = [
        r[0]
        for r in db.query(User.id)
        .filter(
            User.course_id.in_(course_ids),
            User.role == UserRole.student,
            User.is_active == True,  # noqa: E712
        )
        .all()
    ]

    if not student_ids or not subject_ids:
        # Return empty PDF with note
        buf = io.BytesIO()
        doc = _base_doc(buf, "Defaulters Report")
        story = []
        _header_block(
            story,
            "Defaulters Report",
            [
                f"Threshold: {threshold}%  |  Semester: {semester or 'All'}  |  "
                f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}",
            ],
        )
        story.append(Paragraph("No defaulters found for the given filters.", BODY_STYLE))
        doc.build(story)
        buf.seek(0)
        return buf

    ended_session_ids = [
        r[0]
        for r in db.query(AttendanceSession.id)
        .filter(
            AttendanceSession.subject_id.in_(subject_ids),
            AttendanceSession.status == SessionStatus.ended,
        )
        .all()
    ]

    rows = []
    if ended_session_ids:
        subq = (
            db.query(
                AttendanceRecord.student_id,
                AttendanceSession.subject_id,
                func.count(AttendanceRecord.id).label("total"),
                func.sum(
                    case((AttendanceRecord.status == AttendanceStatus.present, 1), else_=0)
                ).label("present"),
            )
            .join(AttendanceSession, AttendanceRecord.session_id == AttendanceSession.id)
            .filter(
                AttendanceRecord.session_id.in_(ended_session_ids),
                AttendanceRecord.student_id.in_(student_ids),
            )
            .group_by(AttendanceRecord.student_id, AttendanceSession.subject_id)
            .subquery()
        )
        rows = (
            db.query(
                subq.c.student_id,
                subq.c.subject_id,
                subq.c.total,
                subq.c.present,
                User.name.label("student_name"),
                User.roll_number,
                Subject.name.label("subject_name"),
            )
            .join(User, User.id == subq.c.student_id)
            .join(Subject, Subject.id == subq.c.subject_id)
            .filter(
                subq.c.total > 0,
                (subq.c.present * 100.0 / subq.c.total) < threshold,
            )
            .order_by(subq.c.present * 100.0 / subq.c.total)
            .all()
        )

    buf = io.BytesIO()
    doc = _base_doc(buf, "Defaulters Report", landscape_mode=True)
    story = []

    _header_block(
        story,
        "Defaulters Report",
        [
            f"Threshold: {threshold}%   |   Semester: {semester or 'All'}   |   "
            f"Total defaulters: {len(rows)}",
            f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}",
        ],
    )

    if not rows:
        story.append(
            Paragraph(
                f"No students below {threshold}% attendance threshold.",
                BODY_STYLE,
            )
        )
    else:
        data = [["#", "Roll No", "Name", "Subject", "Present", "Total", "Attendance %", "Status"]]
        for idx, r in enumerate(rows, 1):
            pct = round(r.present * 100 / r.total, 1) if r.total else 0.0
            st = (
                "DETAINED"
                if pct < threshold - 25
                else "CRITICAL" if pct < threshold - 10 else "WARNING"
            )
            data.append(
                [
                    idx,
                    r.roll_number or "—",
                    r.student_name,
                    r.subject_name,
                    r.present,
                    r.total,
                    f"{pct}%",
                    st,
                ]
            )

        col_widths = [1 * cm, 2.5 * cm, 5 * cm, 5 * cm, 2 * cm, 2 * cm, 3 * cm, 2.5 * cm]
        tbl = Table(data, colWidths=col_widths, repeatRows=1)
        ts = _table_style()
        for i, row in enumerate(data[1:], start=1):
            try:
                pct_val = float(str(row[6]).replace("%", ""))
            except ValueError:
                continue
            c = _status_color(pct_val, threshold)
            ts.add("TEXTCOLOR", (6, i), (6, i), c)
            ts.add("FONTNAME", (6, i), (6, i), "Helvetica-Bold")
            ts.add("TEXTCOLOR", (7, i), (7, i), c)
            ts.add("FONTNAME", (7, i), (7, i), "Helvetica-Bold")
        tbl.setStyle(ts)
        story.append(tbl)

    doc.build(story)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════════════════
# 4. Monthly Excel Matrix
# ═══════════════════════════════════════════════════════════════════════


def generate_monthly_excel(
    subject_id: int,
    year: int,
    month: int,
    db: Session,
) -> io.BytesIO:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    from config import settings
    from database import (
        AttendanceRecord,
        AttendanceSession,
        SessionStatus,
        Subject,
        User,
        UserRole,
    )

    subj = db.query(Subject).filter(Subject.id == subject_id).first()
    if not subj:
        raise ValueError(f"Subject {subject_id} not found.")

    # Sessions in this month
    month_start = date(year, month, 1)
    month_end = date(year, month, calendar.monthrange(year, month)[1])

    sessions = (
        db.query(AttendanceSession)
        .filter(
            AttendanceSession.subject_id == subject_id,
            AttendanceSession.status == SessionStatus.ended,
            AttendanceSession.date >= month_start,
            AttendanceSession.date <= month_end,
        )
        .order_by(AttendanceSession.date, AttendanceSession.start_time)
        .all()
    )

    # Students enrolled (via same course)
    students = (
        db.query(User)
        .filter(
            User.course_id == subj.course_id,
            User.role == UserRole.student,
            User.is_active == True,  # noqa: E712
        )
        .order_by(User.roll_number)
        .all()
    )

    # All records for these sessions
    session_ids = [s.id for s in sessions]
    records_q = (
        (db.query(AttendanceRecord).filter(AttendanceRecord.session_id.in_(session_ids)).all())
        if session_ids
        else []
    )

    # Lookup: {(student_id, session_id): status_value}
    lookup: dict[tuple, str] = {}
    for rec in records_q:
        lookup[(rec.student_id, rec.session_id)] = rec.status.value[0].upper()  # P/A/L/M/D

    # ── Build workbook ────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{subj.code}_{year}_{month:02d}"

    # Styles
    FILL_HEADER = PatternFill("solid", fgColor="1E40AF")
    FILL_GREEN = PatternFill("solid", fgColor="BBF7D0")
    FILL_RED = PatternFill("solid", fgColor="FEE2E2")
    FILL_AMBER = PatternFill("solid", fgColor="FEF3C7")
    FILL_BLUE = PatternFill("solid", fgColor="DBEAFE")
    FILL_PURPLE = PatternFill("solid", fgColor="EDE9FE")
    FILL_TOTAL = PatternFill("solid", fgColor="E2E8F0")
    FONT_HEADER = Font(bold=True, color="FFFFFF", size=9)
    FONT_TITLE = Font(bold=True, color="1E3A5F", size=12)
    FONT_NORMAL = Font(size=8)
    ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
    thin_side = Side(style="thin", color="CBD5E1")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # Title rows
    title_text = f"Monthly Attendance Matrix — {subj.name} ({subj.code})"
    ws["A1"] = title_text
    ws["A1"].font = FONT_TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4 + len(sessions))
    ws["A2"] = f"{calendar.month_name[month]} {year}   |   Total sessions: {len(sessions)}"
    ws["A2"].font = Font(size=9, color="64748B")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4 + len(sessions))
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 14

    # Header row (row 3)
    HDR_ROW = 3
    ws.cell(HDR_ROW, 1, "#").font = FONT_HEADER
    ws.cell(HDR_ROW, 1).fill = FILL_HEADER
    ws.cell(HDR_ROW, 1).alignment = ALIGN_CENTER
    ws.cell(HDR_ROW, 1).border = thin_border
    ws.column_dimensions["A"].width = 4

    ws.cell(HDR_ROW, 2, "Roll No").font = FONT_HEADER
    ws.cell(HDR_ROW, 2).fill = FILL_HEADER
    ws.cell(HDR_ROW, 2).alignment = ALIGN_CENTER
    ws.cell(HDR_ROW, 2).border = thin_border
    ws.column_dimensions["B"].width = 12

    ws.cell(HDR_ROW, 3, "Name").font = FONT_HEADER
    ws.cell(HDR_ROW, 3).fill = FILL_HEADER
    ws.cell(HDR_ROW, 3).alignment = ALIGN_LEFT
    ws.cell(HDR_ROW, 3).border = thin_border
    ws.column_dimensions["C"].width = 22

    for col_idx, sess in enumerate(sessions, start=4):
        cell = ws.cell(
            HDR_ROW, col_idx, f"{sess.date.strftime('%d/%m')}\n{str(sess.start_time)[:5]}"
        )
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = 6

    total_col = 4 + len(sessions)
    ws.cell(HDR_ROW, total_col, "% Att").font = FONT_HEADER
    ws.cell(HDR_ROW, total_col).fill = FILL_HEADER
    ws.cell(HDR_ROW, total_col).alignment = ALIGN_CENTER
    ws.cell(HDR_ROW, total_col).border = thin_border
    ws.column_dimensions[get_column_letter(total_col)].width = 8
    ws.row_dimensions[HDR_ROW].height = 30

    # Data rows
    status_fill = {
        "P": FILL_GREEN,
        "A": FILL_RED,
        "L": FILL_AMBER,
        "M": FILL_BLUE,
        "D": FILL_PURPLE,
    }

    for row_idx, stu in enumerate(students, start=HDR_ROW + 1):
        ws.cell(row_idx, 1, row_idx - HDR_ROW).font = FONT_NORMAL
        ws.cell(row_idx, 1).alignment = ALIGN_CENTER
        ws.cell(row_idx, 1).border = thin_border

        ws.cell(row_idx, 2, stu.roll_number or "—").font = FONT_NORMAL
        ws.cell(row_idx, 2).alignment = ALIGN_CENTER
        ws.cell(row_idx, 2).border = thin_border

        ws.cell(row_idx, 3, stu.name).font = FONT_NORMAL
        ws.cell(row_idx, 3).alignment = ALIGN_LEFT
        ws.cell(row_idx, 3).border = thin_border

        present_count = 0
        for col_idx, sess in enumerate(sessions, start=4):
            val = lookup.get((stu.id, sess.id), "—")
            cell = ws.cell(row_idx, col_idx, val)
            cell.font = Font(bold=True, size=8) if val in status_fill else FONT_NORMAL
            cell.fill = status_fill.get(val, PatternFill())
            cell.alignment = ALIGN_CENTER
            cell.border = thin_border
            if val == "P":
                present_count += 1

        pct = round(present_count * 100 / len(sessions), 1) if sessions else 0.0
        pct_cell = ws.cell(row_idx, total_col, f"{pct}%")
        pct_cell.font = Font(
            bold=True,
            size=8,
            color=(
                "166534"
                if pct >= settings.ATTENDANCE_THRESHOLD
                else ("92400E" if pct >= settings.ATTENDANCE_THRESHOLD - 10 else "991B1B")
            ),
        )
        pct_cell.fill = FILL_TOTAL
        pct_cell.alignment = ALIGN_CENTER
        pct_cell.border = thin_border

    # Legend row
    legend_row = HDR_ROW + len(students) + 2
    ws.cell(legend_row, 1, "Legend:").font = Font(bold=True, size=8)
    items = [
        ("P", "Present", FILL_GREEN),
        ("A", "Absent", FILL_RED),
        ("L", "Late", FILL_AMBER),
        ("M", "Medical Leave", FILL_BLUE),
        ("D", "Duty Leave", FILL_PURPLE),
    ]
    for i, (code, label, fill) in enumerate(items):
        c = ws.cell(legend_row, 2 + i * 2, code)
        c.fill = fill
        c.font = Font(bold=True, size=8)
        c.alignment = ALIGN_CENTER
        ws.cell(legend_row, 3 + i * 2, label).font = Font(size=8)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

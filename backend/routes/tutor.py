"""
AutoAttend AI v2.0 — Tutor Routes

Assignment Management (HOD only):
  POST   /api/tutor/assign                     — bulk assign students to a tutor
  POST   /api/tutor/assign-by-roll-range       — assign by roll number range
  POST   /api/tutor/assign-by-section          — assign all students in a section
  POST   /api/tutor/import-excel               — bulk assign via Excel
  POST   /api/tutor/export-assignments-excel   — export to Excel
  DELETE /api/tutor/remove/{assignment_id}      — remove one assignment
  POST   /api/tutor/deactivate-all             — deactivate all for a year
  GET    /api/tutor/assignments                — list with filters
  GET    /api/tutor/unassigned-students        — students with no tutor

Teacher Tutor Dashboard:
  GET    /api/tutor/my-ward-students           — teacher's ward students + attendance
  GET    /api/tutor/ward-student/{student_id}/full-report
  GET    /api/tutor/my-defaulters              — ward students below threshold
  PATCH  /api/tutor/ward/{student_id}/contacts — edit ward student phone numbers
  POST   /api/tutor/notify-ward               — send notification to ward students
"""

import io
import logging
from datetime import datetime, timedelta, timezone
from typing import List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import func as sqlfunc
from sqlalchemy.orm import Session

from config import settings
from database import (
    AlertChannel,
    AlertsLog,
    AlertStatus,
    AttendanceRecord,
    AttendanceSession,
    AttendanceStatus,
    Section,
    SessionStatus,
    Subject,
    TutorAssignment,
    User,
    UserRole,
    get_db,
)
from utils.auth_utils import hod_or_above, teacher_or_above
from utils.notification_utils import send_push_notification
from utils.sms import send_sms
from utils.whatsapp import send_whatsapp_message

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/tutor", tags=["tutor"])

# ── Threshold constants ────────────────────────────────────────────────
_THRESHOLD = settings.ATTENDANCE_THRESHOLD  # 75.0


def _attendance_label(pct: float) -> str:
    if pct >= _THRESHOLD:
        return "safe"
    if pct >= 60.0:
        return "warning"
    if pct >= 50.0:
        return "critical"
    return "detained"


# ═══════════════════════════════════════════════════════════════════════
# Pydantic schemas
# ═══════════════════════════════════════════════════════════════════════

class AssignRequest(BaseModel):
    tutor_id:      int
    student_ids:   List[int]
    academic_year: str = Field(..., min_length=4, max_length=20)
    note:          Optional[str] = None
    force:         bool = False


class AssignByRollRange(BaseModel):
    tutor_id:      int
    roll_start:    str
    roll_end:      str
    academic_year: str = Field(..., min_length=4, max_length=20)
    note:          Optional[str] = None
    force:         bool = False


class AssignBySection(BaseModel):
    tutor_id:       int
    section_id:     int
    academic_year:  str = Field(..., min_length=4, max_length=20)
    skip_existing:  bool = True
    note:           Optional[str] = None


class DeactivateAllRequest(BaseModel):
    academic_year: str


class NotifyWardRequest(BaseModel):
    student_ids:  List[int]
    message:      str = Field("", max_length=1600)
    channels:     List[str] = ["push"]  # push, whatsapp, sms
    use_template: bool = False  # auto-generate per-student attendance report


class UpdateContactRequest(BaseModel):
    phone:        Optional[str] = None
    parent_phone: Optional[str] = None


# ═══════════════════════════════════════════════════════════════════════
# Helper: compute per-student attendance
# ═══════════════════════════════════════════════════════════════════════

def _student_attendance_summary(student_id: int, db: Session) -> dict:
    """Return overall + per-subject attendance for a student."""
    # All records for this student where session is ended (not expired)
    records = (
        db.query(AttendanceRecord)
        .join(AttendanceSession, AttendanceRecord.session_id == AttendanceSession.id)
        .filter(
            AttendanceRecord.student_id == student_id,
            AttendanceSession.status.in_([SessionStatus.ended, SessionStatus.active]),
        )
        .all()
    )

    if not records:
        return {
            "overall_pct": 0.0,
            "per_subject": [],
            "total_sessions": 0,
            "present_sessions": 0,
        }

    # Aggregate per subject
    subject_map = {}  # subject_id -> {present, total, name}
    for rec in records:
        sess = rec.session
        sid = sess.subject_id
        if sid not in subject_map:
            subj = db.query(Subject).filter(Subject.id == sid).first()
            subject_map[sid] = {
                "subject_id": sid,
                "subject_name": subj.name if subj else f"Subject #{sid}",
                "subject_code": subj.code if subj else "",
                "present": 0,
                "total": 0,
            }
        subject_map[sid]["total"] += 1
        if rec.status in (AttendanceStatus.present, AttendanceStatus.late):
            subject_map[sid]["present"] += 1

    total = sum(s["total"] for s in subject_map.values())
    present = sum(s["present"] for s in subject_map.values())
    overall_pct = round((present / total * 100) if total > 0 else 0.0, 1)

    per_subject = []
    for s in subject_map.values():
        pct = round((s["present"] / s["total"] * 100) if s["total"] > 0 else 0.0, 1)
        per_subject.append({
            "subject_name": s["subject_name"],
            "subject_code": s["subject_code"],
            "present": s["present"],
            "total": s["total"],
            "pct": pct,
        })

    return {
        "overall_pct": overall_pct,
        "per_subject": per_subject,
        "total_sessions": total,
        "present_sessions": present,
    }


# ═══════════════════════════════════════════════════════════════════════
# POST /api/tutor/assign — manual bulk assign
# ═══════════════════════════════════════════════════════════════════════

@router.post("/assign")
def assign_students(
    body:         AssignRequest,
    current_user: dict    = Depends(hod_or_above),
    db:           Session = Depends(get_db),
):
    # Validate tutor is a teacher in the same college
    tutor = db.query(User).filter(
        User.id == body.tutor_id,
        User.role.in_([UserRole.teacher, UserRole.hod, UserRole.principal]),
        User.college_id == current_user["college_id"],
        User.is_active == True,
    ).first()
    if not tutor:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Tutor not found or not a teacher in your college.")

    assigned = 0
    conflicts = []
    not_found = []

    for sid in body.student_ids:
        student = db.query(User).filter(User.id == sid, User.role == UserRole.student).first()
        if not student:
            not_found.append(sid)
            continue

        existing = (
            db.query(TutorAssignment)
            .filter(
                TutorAssignment.student_id == sid,
                TutorAssignment.academic_year == body.academic_year,
                TutorAssignment.is_active == True,
            )
            .first()
        )
        if existing:
            if not body.force:
                ex_tutor = db.query(User).filter(User.id == existing.tutor_id).first()
                conflicts.append({
                    "student_id": sid,
                    "student_name": student.name,
                    "existing_tutor": ex_tutor.name if ex_tutor else "Unknown",
                    "existing_tutor_id": existing.tutor_id,
                })
                continue
            # Force reassign — delete old row to satisfy unique constraint
            db.delete(existing)
            db.flush()

        db.add(TutorAssignment(
            tutor_id=body.tutor_id,
            student_id=sid,
            academic_year=body.academic_year,
            assigned_by=current_user["id"],
            note=body.note,
        ))
        assigned += 1

    db.commit()

    logger.info("🎓 TUTOR ASSIGN │ tutor_id=%d │ assigned=%d │ conflicts=%d │ not_found=%d │ by user_id=%d",
                body.tutor_id, assigned, len(conflicts), len(not_found), current_user["id"])

    return {
        "assigned": assigned,
        "conflicts": conflicts,
        "not_found": not_found,
        "tutor_name": tutor.name,
    }


# ═══════════════════════════════════════════════════════════════════════
# POST /api/tutor/assign-by-roll-range
# ═══════════════════════════════════════════════════════════════════════

@router.post("/assign-by-roll-range")
def assign_by_roll_range(
    body:         AssignByRollRange,
    current_user: dict    = Depends(hod_or_above),
    db:           Session = Depends(get_db),
):
    tutor = db.query(User).filter(
        User.id == body.tutor_id,
        User.role.in_([UserRole.teacher, UserRole.hod, UserRole.principal]),
        User.college_id == current_user["college_id"],
        User.is_active == True,
    ).first()
    if not tutor:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Tutor not found or not a teacher in your college.")

    # Find students whose roll_number is between roll_start and roll_end (inclusive, lexicographic)
    students = (
        db.query(User)
        .filter(
            User.role == UserRole.student,
            User.is_active == True,
            User.roll_number.isnot(None),
            User.roll_number >= body.roll_start.strip(),
            User.roll_number <= body.roll_end.strip(),
            User.college_id == current_user["college_id"],
        )
        .order_by(User.roll_number)
        .all()
    )

    if not students:
        raise HTTPException(status.HTTP_404_NOT_FOUND,
                            f"No students found between {body.roll_start} and {body.roll_end}.")

    # Delegate to the assign logic
    student_ids = [s.id for s in students]
    req = AssignRequest(
        tutor_id=body.tutor_id,
        student_ids=student_ids,
        academic_year=body.academic_year,
        note=body.note,
        force=body.force,
    )
    return assign_students(req, current_user, db)


# ═══════════════════════════════════════════════════════════════════════
# POST /api/tutor/assign-by-section
# ═══════════════════════════════════════════════════════════════════════

@router.post("/assign-by-section")
def assign_by_section(
    body:         AssignBySection,
    current_user: dict    = Depends(hod_or_above),
    db:           Session = Depends(get_db),
):
    section = db.query(Section).filter(Section.id == body.section_id).first()
    if not section:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Section not found.")

    tutor = db.query(User).filter(
        User.id == body.tutor_id,
        User.role.in_([UserRole.teacher, UserRole.hod, UserRole.principal]),
        User.college_id == current_user["college_id"],
        User.is_active == True,
    ).first()
    if not tutor:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Tutor not found.")

    students = (
        db.query(User)
        .filter(
            User.section_id == body.section_id,
            User.role == UserRole.student,
            User.is_active == True,
        )
        .all()
    )

    assigned = 0
    skipped = 0

    for student in students:
        existing = (
            db.query(TutorAssignment)
            .filter(
                TutorAssignment.student_id == student.id,
                TutorAssignment.academic_year == body.academic_year,
                TutorAssignment.is_active == True,
            )
            .first()
        )
        if existing:
            if body.skip_existing:
                skipped += 1
                continue
            # Delete old to satisfy unique constraint on (student_id, academic_year)
            db.delete(existing)
            db.flush()

        db.add(TutorAssignment(
            tutor_id=body.tutor_id,
            student_id=student.id,
            academic_year=body.academic_year,
            assigned_by=current_user["id"],
            note=body.note,
        ))
        assigned += 1

    db.commit()

    logger.info("🎓 TUTOR ASSIGN-BY-SECTION │ section=%s │ tutor=%s │ assigned=%d │ skipped=%d",
                section.name, tutor.name, assigned, skipped)

    return {
        "assigned": assigned,
        "skipped": skipped,
        "section_name": section.name,
        "tutor_name": tutor.name,
    }


# ═══════════════════════════════════════════════════════════════════════
# POST /api/tutor/import-excel
# ═══════════════════════════════════════════════════════════════════════

@router.post("/import-excel")
def import_excel(
    file:         UploadFile = File(...),
    current_user: dict    = Depends(hod_or_above),
    db:           Session = Depends(get_db),
):
    """
    Excel columns: roll_number, tutor_email, academic_year
    """
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Please upload an Excel file (.xlsx).")

    import openpyxl

    try:
        wb = openpyxl.load_workbook(file.file, read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Invalid Excel file.")

    ws = wb.active
    if ws is None:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "No active sheet.")

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Excel must have header + data rows.")

    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    required = {"roll_number", "tutor_email", "academic_year"}
    missing = required - set(header)
    if missing:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, f"Missing columns: {missing}")

    rn_idx = header.index("roll_number")
    te_idx = header.index("tutor_email")
    ay_idx = header.index("academic_year")

    success = 0
    skipped = 0
    failed_rows = []

    for row_num, row in enumerate(rows[1:], start=2):
        roll = str(row[rn_idx]).strip() if rn_idx < len(row) and row[rn_idx] else None
        email = str(row[te_idx]).strip() if te_idx < len(row) and row[te_idx] else None
        year = str(row[ay_idx]).strip() if ay_idx < len(row) and row[ay_idx] else None

        if not roll or not email or not year:
            failed_rows.append({"row": row_num, "reason": "Missing data"})
            continue

        student = db.query(User).filter(User.roll_number == roll, User.role == UserRole.student).first()
        if not student:
            failed_rows.append({"row": row_num, "reason": f"Student '{roll}' not found"})
            continue

        tutor = db.query(User).filter(
            User.email == email,
            User.role.in_([UserRole.teacher, UserRole.hod, UserRole.principal]),
        ).first()
        if not tutor:
            failed_rows.append({"row": row_num, "reason": f"Tutor '{email}' not found"})
            continue

        existing = (
            db.query(TutorAssignment)
            .filter(
                TutorAssignment.student_id == student.id,
                TutorAssignment.academic_year == year,
                TutorAssignment.is_active == True,
            )
            .first()
        )
        if existing:
            skipped += 1
            continue

        db.add(TutorAssignment(
            tutor_id=tutor.id,
            student_id=student.id,
            academic_year=year,
            assigned_by=current_user["id"],
        ))
        success += 1

    db.commit()
    wb.close()

    logger.info("🎓 TUTOR IMPORT-EXCEL │ success=%d │ skipped=%d │ failed=%d │ by user_id=%d",
                success, skipped, len(failed_rows), current_user["id"])

    return {
        "total": len(rows) - 1,
        "success_count": success,
        "skipped_count": skipped,
        "failed_rows": failed_rows,
    }


# ═══════════════════════════════════════════════════════════════════════
# POST /api/tutor/export-assignments-excel
# ═══════════════════════════════════════════════════════════════════════

@router.post("/export-assignments-excel")
def export_assignments_excel(
    academic_year: str,
    current_user:  dict    = Depends(hod_or_above),
    db:            Session = Depends(get_db),
):
    import openpyxl

    assignments = (
        db.query(TutorAssignment)
        .filter(
            TutorAssignment.academic_year == academic_year,
            TutorAssignment.is_active == True,
        )
        .order_by(TutorAssignment.tutor_id)
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tutor Assignments"
    ws.append([
        "Tutor Name", "Tutor Email", "Student Roll No", "Student Name",
        "Section", "Academic Year", "Assigned Date",
    ])

    for a in assignments:
        tutor = db.query(User).filter(User.id == a.tutor_id).first()
        student = db.query(User).filter(User.id == a.student_id).first()
        section_name = ""
        if student and student.section_id:
            sec = db.query(Section).filter(Section.id == student.section_id).first()
            section_name = sec.name if sec else ""

        ws.append([
            tutor.name if tutor else "",
            tutor.email if tutor else "",
            student.roll_number if student else "",
            student.name if student else "",
            section_name,
            a.academic_year,
            a.assigned_at.strftime("%Y-%m-%d %H:%M") if a.assigned_at else "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    wb.close()

    logger.info("🎓 TUTOR EXPORT │ year=%s │ rows=%d │ by user_id=%d",
                academic_year, len(assignments), current_user["id"])

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=tutor_assignments_{academic_year}.xlsx"},
    )


# ═══════════════════════════════════════════════════════════════════════
# DELETE /api/tutor/remove/{assignment_id}
# ═══════════════════════════════════════════════════════════════════════

@router.delete("/remove/{assignment_id}", status_code=status.HTTP_204_NO_CONTENT)
def remove_assignment(
    assignment_id: int,
    current_user:  dict    = Depends(hod_or_above),
    db:            Session = Depends(get_db),
):
    a = db.query(TutorAssignment).filter(TutorAssignment.id == assignment_id).first()
    if not a:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Assignment not found.")
    db.delete(a)
    db.commit()
    logger.info("🎓 TUTOR REMOVE │ assignment_id=%d │ by user_id=%d", assignment_id, current_user["id"])


# ═══════════════════════════════════════════════════════════════════════
# POST /api/tutor/deactivate-all
# ═══════════════════════════════════════════════════════════════════════

@router.post("/deactivate-all")
def deactivate_all(
    body:         DeactivateAllRequest,
    current_user: dict    = Depends(hod_or_above),
    db:           Session = Depends(get_db),
):
    # Scope to college: only deactivate assignments whose tutor belongs to the same college
    college_id = current_user["college_id"]
    matching_ids = (
        db.query(TutorAssignment.id)
        .join(User, TutorAssignment.tutor_id == User.id)
        .filter(
            TutorAssignment.academic_year == body.academic_year,
            TutorAssignment.is_active == True,
            User.college_id == college_id,
        )
        .all()
    )
    ids = [row[0] for row in matching_ids]
    count = 0
    if ids:
        count = (
            db.query(TutorAssignment)
            .filter(TutorAssignment.id.in_(ids))
            .update({"is_active": False}, synchronize_session="fetch")
        )
    db.commit()
    logger.info("🎓 TUTOR DEACTIVATE-ALL │ year=%s │ college_id=%s │ count=%d │ by user_id=%d",
                body.academic_year, college_id, count, current_user["id"])
    return {"deactivated": count, "academic_year": body.academic_year}


# ═══════════════════════════════════════════════════════════════════════
# GET /api/tutor/assignments — list with filters
# ═══════════════════════════════════════════════════════════════════════

@router.get("/assignments")
def list_assignments(
    tutor_id:      Optional[int] = None,
    academic_year: Optional[str] = None,
    section_id:    Optional[int] = None,
    current_user:  dict    = Depends(hod_or_above),
    db:            Session = Depends(get_db),
):
    q = db.query(TutorAssignment).filter(TutorAssignment.is_active == True)

    if tutor_id is not None:
        q = q.filter(TutorAssignment.tutor_id == tutor_id)
    if academic_year:
        q = q.filter(TutorAssignment.academic_year == academic_year)

    assignments = q.order_by(TutorAssignment.tutor_id, TutorAssignment.student_id).all()

    result = []
    for a in assignments:
        tutor = db.query(User).filter(User.id == a.tutor_id).first()
        student = db.query(User).filter(User.id == a.student_id).first()
        section_name = ""
        student_section_id = None
        if student and student.section_id:
            sec = db.query(Section).filter(Section.id == student.section_id).first()
            section_name = sec.name if sec else ""
            student_section_id = student.section_id

        # Apply section filter in-memory (student's section)
        if section_id is not None and student_section_id != section_id:
            continue

        result.append({
            "id": a.id,
            "tutor_id": a.tutor_id,
            "tutor_name": tutor.name if tutor else "",
            "tutor_email": tutor.email if tutor else "",
            "student_id": a.student_id,
            "student_name": student.name if student else "",
            "student_roll": student.roll_number if student else "",
            "section_name": section_name,
            "academic_year": a.academic_year,
            "note": a.note,
            "assigned_at": a.assigned_at.isoformat() if a.assigned_at else None,
        })

    return result


# ═══════════════════════════════════════════════════════════════════════
# GET /api/tutor/unassigned-students
# ═══════════════════════════════════════════════════════════════════════

@router.get("/unassigned-students")
def unassigned_students(
    academic_year: str,
    section_id:    Optional[int] = None,
    current_user:  dict    = Depends(hod_or_above),
    db:            Session = Depends(get_db),
):
    # All active students in college
    q = db.query(User).filter(
        User.role == UserRole.student,
        User.is_active == True,
        User.college_id == current_user["college_id"],
    )
    if section_id is not None:
        q = q.filter(User.section_id == section_id)

    students = q.order_by(User.roll_number).all()

    # IDs of students who already have an active tutor for this year
    assigned_ids = set(
        row[0] for row in
        db.query(TutorAssignment.student_id)
        .filter(
            TutorAssignment.academic_year == academic_year,
            TutorAssignment.is_active == True,
        )
        .all()
    )

    result = []
    for s in students:
        if s.id not in assigned_ids:
            sec_name = ""
            if s.section_id:
                sec = db.query(Section).filter(Section.id == s.section_id).first()
                sec_name = sec.name if sec else ""
            result.append({
                "id": s.id,
                "name": s.name,
                "roll_number": s.roll_number,
                "email": s.email,
                "semester": s.semester,
                "section_name": sec_name,
            })

    return result


# ═══════════════════════════════════════════════════════════════════════
# GET /api/tutor/my-ward-students — teacher dashboard
# ═══════════════════════════════════════════════════════════════════════

@router.get("/my-ward-students")
def my_ward_students(
    academic_year: Optional[str] = None,
    current_user:  dict    = Depends(teacher_or_above),
    db:            Session = Depends(get_db),
):
    year = academic_year or _current_academic_year()

    assignments = (
        db.query(TutorAssignment)
        .filter(
            TutorAssignment.tutor_id == current_user["id"],
            TutorAssignment.academic_year == year,
            TutorAssignment.is_active == True,
        )
        .all()
    )

    result = []
    for a in assignments:
        student = db.query(User).filter(User.id == a.student_id).first()
        if not student:
            continue

        att = _student_attendance_summary(student.id, db)
        sec_name = ""
        if student.section_id:
            sec = db.query(Section).filter(Section.id == student.section_id).first()
            sec_name = sec.name if sec else ""

        pct = att["overall_pct"]
        result.append({
            "student_id": student.id,
            "name": student.name,
            "roll_number": student.roll_number,
            "email": student.email,
            "phone": student.phone or "",
            "parent_phone": student.parent_phone or "",
            "section_name": sec_name,
            "semester": student.semester,
            "academic_year": year,
            "overall_attendance_pct": pct,
            "per_subject": att["per_subject"],
            "attendance_label": _attendance_label(pct),
            "needs_attention": pct < _THRESHOLD,
            "total_sessions": att["total_sessions"],
            "present_sessions": att["present_sessions"],
        })

    logger.info("🎓 TUTOR MY-WARDS │ tutor_id=%d │ year=%s │ count=%d",
                current_user["id"], year, len(result))

    return result


# ═══════════════════════════════════════════════════════════════════════
# GET /api/tutor/ward-student/{student_id}/full-report
# ═══════════════════════════════════════════════════════════════════════

@router.get("/ward-student/{student_id}/full-report")
def ward_student_full_report(
    student_id:    int,
    academic_year: Optional[str] = None,
    current_user:  dict    = Depends(teacher_or_above),
    db:            Session = Depends(get_db),
):
    year = academic_year or _current_academic_year()

    # Verify tutor is assigned to this student
    assignment = (
        db.query(TutorAssignment)
        .filter(
            TutorAssignment.tutor_id == current_user["id"],
            TutorAssignment.student_id == student_id,
            TutorAssignment.academic_year == year,
            TutorAssignment.is_active == True,
        )
        .first()
    )
    if not assignment:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "You are not assigned as tutor to this student.")

    student = db.query(User).filter(User.id == student_id).first()
    if not student:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Student not found.")

    att = _student_attendance_summary(student_id, db)

    # Session-by-session history (last 30 days)
    thirty_days_ago = datetime.now(tz=timezone.utc) - timedelta(days=30)
    recent_records = (
        db.query(AttendanceRecord)
        .join(AttendanceSession, AttendanceRecord.session_id == AttendanceSession.id)
        .filter(
            AttendanceRecord.student_id == student_id,
            AttendanceSession.date >= thirty_days_ago.date(),
        )
        .order_by(AttendanceSession.date.desc())
        .all()
    )

    session_history = []
    for rec in recent_records:
        sess = rec.session
        subj = db.query(Subject).filter(Subject.id == sess.subject_id).first()
        session_history.append({
            "date": sess.date.isoformat(),
            "subject_name": subj.name if subj else "",
            "status": rec.status.value,
            "marked_at": rec.marked_at.isoformat() if rec.marked_at else None,
            "face_verified": rec.face_verified,
            "gps_verified": rec.gps_verified,
        })

    sec_name = ""
    if student.section_id:
        sec = db.query(Section).filter(Section.id == student.section_id).first()
        sec_name = sec.name if sec else ""

    pct = att["overall_pct"]

    return {
        "student_id": student.id,
        "name": student.name,
        "roll_number": student.roll_number,
        "email": student.email,
        "section_name": sec_name,
        "semester": student.semester,
        "overall_attendance_pct": pct,
        "attendance_label": _attendance_label(pct),
        "per_subject": att["per_subject"],
        "session_history_30d": session_history,
    }


# ═══════════════════════════════════════════════════════════════════════
# GET /api/tutor/my-defaulters
# ═══════════════════════════════════════════════════════════════════════

@router.get("/my-defaulters")
def my_defaulters(
    academic_year: Optional[str] = None,
    current_user:  dict    = Depends(teacher_or_above),
    db:            Session = Depends(get_db),
):
    all_wards = my_ward_students(academic_year, current_user, db)
    defaulters = [w for w in all_wards if w["needs_attention"]]
    defaulters.sort(key=lambda x: x["overall_attendance_pct"])

    logger.info("🎓 TUTOR DEFAULTERS │ tutor_id=%d │ count=%d", current_user["id"], len(defaulters))
    return defaulters


# ═══════════════════════════════════════════════════════════════════════
# PATCH /api/tutor/ward/{student_id}/contacts — edit ward student phones
# ═══════════════════════════════════════════════════════════════════════

@router.patch("/ward/{student_id}/contacts")
def update_ward_contacts(
    student_id:   int,
    body:         UpdateContactRequest,
    current_user: dict    = Depends(teacher_or_above),
    db:           Session = Depends(get_db),
):
    year = _current_academic_year()

    # Verify tutor assignment
    assignment = (
        db.query(TutorAssignment)
        .filter(
            TutorAssignment.tutor_id == current_user["id"],
            TutorAssignment.student_id == student_id,
            TutorAssignment.academic_year == year,
            TutorAssignment.is_active == True,
        )
        .first()
    )
    if not assignment:
        raise HTTPException(status.HTTP_403_FORBIDDEN,
                            "You are not assigned as tutor to this student.")

    student = db.query(User).filter(User.id == student_id).first()
    if not student:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Student not found.")

    if body.phone is not None:
        student.phone = body.phone.strip() if body.phone else None
    if body.parent_phone is not None:
        student.parent_phone = body.parent_phone.strip() if body.parent_phone else None

    db.commit()
    db.refresh(student)

    logger.info("🎓 TUTOR CONTACTS │ tutor_id=%d │ student_id=%d │ phone=%s │ parent_phone=%s",
                current_user["id"], student_id, student.phone, student.parent_phone)

    return {
        "student_id": student.id,
        "phone": student.phone or "",
        "parent_phone": student.parent_phone or "",
    }


# ═══════════════════════════════════════════════════════════════════════
# POST /api/tutor/notify-ward — send notifications
# ═══════════════════════════════════════════════════════════════════════

def _build_attendance_template(student: User, att: dict, tutor_name: str) -> str:
    """Build a templated attendance report message for a student."""
    lines = [
        f"Dear Parent/Guardian,",
        f"",
        f"Attendance report for your ward:",
        f"Name: {student.name}",
        f"Roll No: {student.roll_number or '—'}",
        f"",
        f"📊 Per-Subject Breakdown:",
    ]
    for s in att["per_subject"]:
        lines.append(f"  • {s['subject_name']} ({s['subject_code']}): {s['present']}/{s['total']} — {s['pct']}%")

    lines.append(f"")
    lines.append(f"Overall: {att['present_sessions']}/{att['total_sessions']} sessions — {att['overall_pct']}%")

    if att["overall_pct"] < _THRESHOLD:
        lines.append(f"")
        lines.append(f"⚠️ Attendance is below the required {_THRESHOLD}% threshold. Immediate attention is needed.")

    lines.append(f"")
    lines.append(f"— {tutor_name}, Tutor")
    return "\n".join(lines)


@router.post("/notify-ward")
def notify_ward(
    body:         NotifyWardRequest,
    current_user: dict    = Depends(teacher_or_above),
    db:           Session = Depends(get_db),
):
    year = _current_academic_year()

    # Validate all students are wards of this tutor
    ward_ids = set(
        row[0] for row in
        db.query(TutorAssignment.student_id)
        .filter(
            TutorAssignment.tutor_id == current_user["id"],
            TutorAssignment.academic_year == year,
            TutorAssignment.is_active == True,
        )
        .all()
    )

    invalid = [sid for sid in body.student_ids if sid not in ward_ids]
    if invalid:
        raise HTTPException(status.HTTP_403_FORBIDDEN,
                            f"Students {invalid} are not your wards.")

    if not body.use_template and not body.message.strip():
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY,
                            "Message is required when not using template.")

    sent_count = 0
    failed_count = 0

    for sid in body.student_ids:
        student = db.query(User).filter(User.id == sid).first()
        if not student:
            continue

        # Build message: template or custom
        if body.use_template:
            att = _student_attendance_summary(sid, db)
            msg = _build_attendance_template(student, att, current_user["name"])
        else:
            msg = body.message

        for channel in body.channels:
            ok = False
            external_id = None

            if channel == "push":
                ok = send_push_notification(
                    user_id=sid,
                    title="Message from your Tutor",
                    body=msg,
                    db=db,
                )

            elif channel == "whatsapp":
                # Send to BOTH student phone and parent phone
                phones = set()
                if student.phone:
                    phones.add(student.phone.strip())
                if student.parent_phone:
                    phones.add(student.parent_phone.strip())

                for phone in phones:
                    wa = send_whatsapp_message(phone, msg)
                    wa_ok = wa.get("ok", False)
                    db.add(AlertsLog(
                        student_id=sid,
                        alert_type="tutor_notification",
                        message=msg,
                        status=AlertStatus.sent if wa_ok else AlertStatus.failed,
                        channel=AlertChannel.whatsapp,
                        external_id=wa.get("sid"),
                    ))
                    if wa_ok:
                        sent_count += 1
                    else:
                        failed_count += 1
                continue  # already logged per-phone above

            elif channel == "sms":
                # Send SMS via MSG91 to BOTH student phone and parent phone
                phones = set()
                if student.phone:
                    phones.add(student.phone.strip())
                if student.parent_phone:
                    phones.add(student.parent_phone.strip())

                if body.use_template:
                    att_data = att if body.use_template else None
                    if not att_data:
                        att_data = _student_attendance_summary(sid, db)
                    low_subjs = [s["subject_name"] for s in att_data["per_subject"] if s["pct"] < _THRESHOLD]
                    sms_vars = {
                        "student_name": student.name,
                        "roll_no": student.roll_number or "",
                        "attendance_pct": str(att_data["overall_pct"]),
                        "low_subjects": ", ".join(low_subjs) if low_subjs else "None",
                        "tutor_name": current_user["name"],
                    }
                else:
                    sms_vars = {"message": msg}

                for phone in phones:
                    sms_result = send_sms(phone, sms_vars)
                    sms_ok = sms_result.get("ok", False)
                    db.add(AlertsLog(
                        student_id=sid,
                        alert_type="tutor_notification",
                        message=msg[:500],
                        status=AlertStatus.sent if sms_ok else AlertStatus.failed,
                        channel=AlertChannel.sms,
                    ))
                    if sms_ok:
                        sent_count += 1
                    else:
                        failed_count += 1
                continue  # already logged per-phone above

            alert_channel = {
                "push": AlertChannel.email,   # no push enum, use email as proxy
                "whatsapp": AlertChannel.whatsapp,
                "sms": AlertChannel.sms,
            }.get(channel, AlertChannel.email)

            db.add(AlertsLog(
                student_id=sid,
                alert_type="tutor_notification",
                message=msg,
                status=AlertStatus.sent if ok else AlertStatus.failed,
                channel=alert_channel,
                external_id=external_id,
            ))

            if ok:
                sent_count += 1
            else:
                failed_count += 1

    db.commit()

    logger.info("🎓 TUTOR NOTIFY │ tutor_id=%d │ students=%d │ sent=%d │ failed=%d",
                current_user["id"], len(body.student_ids), sent_count, failed_count)

    return {
        "sent": sent_count,
        "failed": failed_count,
        "total_students": len(body.student_ids),
    }


# ═══════════════════════════════════════════════════════════════════════
# Utility
# ═══════════════════════════════════════════════════════════════════════

def _current_academic_year() -> str:
    """Return the current academic year string, e.g. '2025-26'."""
    now = datetime.now(tz=timezone.utc)
    y = now.year
    # Academic year starts in June/July typically
    if now.month < 6:
        return f"{y - 1}-{str(y)[-2:]}"
    return f"{y}-{str(y + 1)[-2:]}"

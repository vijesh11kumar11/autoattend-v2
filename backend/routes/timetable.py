"""
AutoAttend AI v2.0 — Timetable Routes

HOD/Principal Management:
  POST   /api/timetable/entry                  — create entry
  PUT    /api/timetable/entry/{entry_id}       — update entry
  DELETE /api/timetable/entry/{entry_id}       — delete entry
  GET    /api/timetable/department              — full department grid for a day
  POST   /api/timetable/bulk-upload-excel       — upload timetable via Excel
  GET    /api/timetable/export-excel            — export department timetable

Teacher Endpoints:
  GET    /api/timetable/my-today               — today's schedule
  GET    /api/timetable/my-week                — full week schedule
  GET    /api/timetable/my-current-class       — class happening right now
  POST   /api/timetable/start-from-timetable/{timetable_id} — one-tap attendance

Student Endpoint:
  GET    /api/timetable/my-section-timetable   — student's section week timetable
"""

import io
import logging
import secrets
from datetime import date, datetime, timezone
from typing import List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from config import settings
from database import (
    AttendanceRecord,
    AttendanceSession,
    AttendanceStatus,
    Course,
    DayOfWeek,
    Department,
    MarkedBy,
    Section,
    SessionStatus,
    Subject,
    Timetable,
    User,
    UserRole,
    get_db,
)
from utils.auth_utils import any_authenticated, hod_or_above, student_only, teacher_or_above
from utils.bluetooth_utils import generate_bluetooth_token
from utils.notification_utils import send_push_to_many

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/timetable", tags=["timetable"])

DAY_ORDER = ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday"]
DAY_NAMES = {d: d.capitalize() for d in DAY_ORDER}


# ══════════════════════════════════════════════════════════════════════
# Pydantic schemas
# ══════════════════════════════════════════════════════════════════════

class TimetableEntryCreate(BaseModel):
    subject_id:    int
    teacher_id:    int
    day_of_week:   str = Field(..., description="monday..saturday")
    start_time:    str = Field(..., pattern=r"^\d{2}:\d{2}$")
    end_time:      str = Field(..., pattern=r"^\d{2}:\d{2}$")
    room:          Optional[str] = None
    section_id:    Optional[int] = None
    period_number: Optional[int] = None
    is_lab:        bool = False
    color_tag:     Optional[str] = None


class TimetableEntryUpdate(BaseModel):
    subject_id:    Optional[int] = None
    teacher_id:    Optional[int] = None
    day_of_week:   Optional[str] = None
    start_time:    Optional[str] = None
    end_time:      Optional[str] = None
    room:          Optional[str] = None
    section_id:    Optional[int] = None
    period_number: Optional[int] = None
    is_lab:        Optional[bool] = None
    color_tag:     Optional[str] = None


class StartFromTimetableRequest(BaseModel):
    teacher_latitude:  float = Field(..., ge=-90, le=90)
    teacher_longitude: float = Field(..., ge=-180, le=180)


# ══════════════════════════════════════════════════════════════════════
# Helpers
# ══════════════════════════════════════════════════════════════════════

def _validate_day(day: str) -> DayOfWeek:
    try:
        return DayOfWeek(day.lower().strip())
    except ValueError:
        raise HTTPException(status.HTTP_400_BAD_REQUEST,
                            f"Invalid day_of_week '{day}'. Use monday..saturday.")


def _check_teacher_clash(db: Session, teacher_id: int, day: DayOfWeek,
                         start: str, end: str, exclude_id: Optional[int] = None):
    q = db.query(Timetable).filter(
        Timetable.teacher_id == teacher_id,
        Timetable.day_of_week == day,
        Timetable.start_time < end,
        Timetable.end_time > start,
    )
    if exclude_id:
        q = q.filter(Timetable.id != exclude_id)
    clash = q.first()
    if clash:
        subj = db.query(Subject).filter(Subject.id == clash.subject_id).first()
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            f"Teacher clash: already has {subj.name if subj else 'a class'} "
            f"on {day.value} {clash.start_time}–{clash.end_time}.",
        )


def _check_section_clash(db: Session, section_id: int, day: DayOfWeek,
                          start: str, end: str, exclude_id: Optional[int] = None):
    q = db.query(Timetable).filter(
        Timetable.section_id == section_id,
        Timetable.day_of_week == day,
        Timetable.start_time < end,
        Timetable.end_time > start,
    )
    if exclude_id:
        q = q.filter(Timetable.id != exclude_id)
    clash = q.first()
    if clash:
        subj = db.query(Subject).filter(Subject.id == clash.subject_id).first()
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            f"Section clash: section already has {subj.name if subj else 'a class'} "
            f"on {day.value} {clash.start_time}–{clash.end_time}.",
        )


def _today_day() -> DayOfWeek:
    """Return the DayOfWeek enum for today."""
    day_name = datetime.now(tz=timezone.utc).strftime("%A").lower()
    try:
        return DayOfWeek(day_name)
    except ValueError:
        # Sunday — not in enum
        return None


def _session_status_for_entry(entry: Timetable, db: Session) -> dict:
    """Check if there's an attendance session for this timetable entry today."""
    today = date.today()
    sess = (
        db.query(AttendanceSession)
        .filter(
            AttendanceSession.subject_id == entry.subject_id,
            AttendanceSession.teacher_id == entry.teacher_id,
            AttendanceSession.date == today,
        )
    )
    if entry.section_id:
        sess = sess.filter(AttendanceSession.section_id == entry.section_id)

    session = sess.first()
    if not session:
        return {"session_status": "not_started", "session_id": None}
    return {
        "session_status": session.status.value,
        "session_id": session.id,
    }


def _serialize_entry(entry: Timetable, db: Session, include_session: bool = False) -> dict:
    subj = db.query(Subject).filter(Subject.id == entry.subject_id).first() if entry.subject_id else None
    teacher = db.query(User).filter(User.id == entry.teacher_id).first()
    sec_name = ""
    if entry.section_id:
        sec = db.query(Section).filter(Section.id == entry.section_id).first()
        sec_name = sec.name if sec else ""

    result = {
        "timetable_id":  entry.id,
        "subject_id":    entry.subject_id,
        "subject_name":  subj.name if subj else ("TWM" if entry.is_twm else ""),
        "subject_code":  subj.code if subj else ("TWM" if entry.is_twm else ""),
        "teacher_id":    entry.teacher_id,
        "teacher_name":  teacher.name if teacher else "",
        "day_of_week":   entry.day_of_week.value,
        "start_time":    entry.start_time,
        "end_time":      entry.end_time,
        "room":          entry.room or "",
        "section_id":    entry.section_id,
        "section_name":  sec_name,
        "period_number": entry.period_number,
        "is_lab":        entry.is_lab,
        "is_twm":        entry.is_twm or False,
        "color_tag":     entry.color_tag,
    }
    if include_session:
        result.update(_session_status_for_entry(entry, db))
    return result


# ══════════════════════════════════════════════════════════════════════
# POST /api/timetable/entry — create
# ══════════════════════════════════════════════════════════════════════

@router.post("/entry", status_code=status.HTTP_201_CREATED)
def create_entry(
    body:         TimetableEntryCreate,
    current_user: dict    = Depends(hod_or_above),
    db:           Session = Depends(get_db),
):
    day = _validate_day(body.day_of_week)

    # Validate teacher belongs to same college
    teacher = db.query(User).filter(
        User.id == body.teacher_id,
        User.role.in_([UserRole.teacher, UserRole.hod, UserRole.principal]),
        User.college_id == current_user["college_id"],
        User.is_active == True,
    ).first()
    if not teacher:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Teacher not found in your college.")

    # Validate subject exists
    subj = db.query(Subject).filter(Subject.id == body.subject_id).first()
    if not subj:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Subject not found.")

    # Clash checks
    _check_teacher_clash(db, body.teacher_id, day, body.start_time, body.end_time)
    if body.section_id:
        _check_section_clash(db, body.section_id, day, body.start_time, body.end_time)

    entry = Timetable(
        subject_id=body.subject_id,
        teacher_id=body.teacher_id,
        day_of_week=day,
        start_time=body.start_time,
        end_time=body.end_time,
        room=body.room,
        section_id=body.section_id,
        period_number=body.period_number,
        is_lab=body.is_lab,
        color_tag=body.color_tag,
    )
    db.add(entry)
    db.commit()
    db.refresh(entry)

    logger.info("📅 TIMETABLE CREATE │ id=%d │ %s %s–%s │ subj=%s │ teacher=%s │ by user_id=%d",
                entry.id, day.value, body.start_time, body.end_time,
                subj.name, teacher.name, current_user["id"])

    return _serialize_entry(entry, db)


# ══════════════════════════════════════════════════════════════════════
# PUT /api/timetable/entry/{entry_id} — update
# ══════════════════════════════════════════════════════════════════════

@router.put("/entry/{entry_id}")
def update_entry(
    entry_id:     int,
    body:         TimetableEntryUpdate,
    current_user: dict    = Depends(hod_or_above),
    db:           Session = Depends(get_db),
):
    entry = db.query(Timetable).filter(Timetable.id == entry_id).first()
    if not entry:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Timetable entry not found.")

    # Apply updates
    if body.subject_id is not None:
        subj = db.query(Subject).filter(Subject.id == body.subject_id).first()
        if not subj:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Subject not found.")
        entry.subject_id = body.subject_id

    if body.teacher_id is not None:
        teacher = db.query(User).filter(
            User.id == body.teacher_id,
            User.college_id == current_user["college_id"],
        ).first()
        if not teacher:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Teacher not found.")
        entry.teacher_id = body.teacher_id

    if body.day_of_week is not None:
        entry.day_of_week = _validate_day(body.day_of_week)
    if body.start_time is not None:
        entry.start_time = body.start_time
    if body.end_time is not None:
        entry.end_time = body.end_time
    if body.room is not None:
        entry.room = body.room
    if body.section_id is not None:
        entry.section_id = body.section_id
    if body.period_number is not None:
        entry.period_number = body.period_number
    if body.is_lab is not None:
        entry.is_lab = body.is_lab
    if body.color_tag is not None:
        entry.color_tag = body.color_tag

    # Clash checks with updated values
    _check_teacher_clash(db, entry.teacher_id, entry.day_of_week,
                         entry.start_time, entry.end_time, exclude_id=entry.id)
    if entry.section_id:
        _check_section_clash(db, entry.section_id, entry.day_of_week,
                             entry.start_time, entry.end_time, exclude_id=entry.id)

    db.commit()
    db.refresh(entry)

    logger.info("📅 TIMETABLE UPDATE │ id=%d │ by user_id=%d", entry.id, current_user["id"])
    return _serialize_entry(entry, db)


# ══════════════════════════════════════════════════════════════════════
# DELETE /api/timetable/entry/{entry_id}
# ══════════════════════════════════════════════════════════════════════

@router.delete("/entry/{entry_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_entry(
    entry_id:     int,
    current_user: dict    = Depends(hod_or_above),
    db:           Session = Depends(get_db),
):
    entry = db.query(Timetable).filter(Timetable.id == entry_id).first()
    if not entry:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Timetable entry not found.")
    db.delete(entry)
    db.commit()
    logger.info("📅 TIMETABLE DELETE │ id=%d │ by user_id=%d", entry_id, current_user["id"])


# ══════════════════════════════════════════════════════════════════════
# GET /api/timetable/department — full department grid for a day
# ══════════════════════════════════════════════════════════════════════

@router.get("/department")
def department_timetable(
    department_id: Optional[int] = None,
    day_of_week:   Optional[str] = None,
    current_user:  dict    = Depends(hod_or_above),
    db:            Session = Depends(get_db),
):
    dept_id = department_id or current_user.get("department_id")
    if not dept_id:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "department_id required.")

    course_ids = [r[0] for r in db.query(Course.id).filter(Course.department_id == dept_id).all()]
    if not course_ids:
        return {"entries": []}

    subject_ids = [r[0] for r in db.query(Subject.id).filter(Subject.course_id.in_(course_ids)).all()]
    if not subject_ids:
        return {"entries": []}

    q = db.query(Timetable).filter(Timetable.subject_id.in_(subject_ids))
    if day_of_week:
        q = q.filter(Timetable.day_of_week == _validate_day(day_of_week))

    entries = q.order_by(Timetable.day_of_week, Timetable.start_time).all()

    # Group by day
    by_day = {d: [] for d in DAY_ORDER}
    for e in entries:
        key = e.day_of_week.value
        if key in by_day:
            by_day[key].append(_serialize_entry(e, db))

    if day_of_week:
        d = day_of_week.lower().strip()
        return {"day": DAY_NAMES.get(d, d), "entries": by_day.get(d, [])}

    return {
        "timetable": [
            {"day": DAY_NAMES[d], "entries": by_day[d]}
            for d in DAY_ORDER if by_day[d]
        ]
    }


# ══════════════════════════════════════════════════════════════════════
# POST /api/timetable/bulk-upload-excel
# ══════════════════════════════════════════════════════════════════════

@router.post("/bulk-upload-excel")
def bulk_upload_excel(
    file:         UploadFile = File(...),
    current_user: dict    = Depends(hod_or_above),
    db:           Session = Depends(get_db),
):
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Please upload an Excel file (.xlsx).")

    import openpyxl

    try:
        wb = openpyxl.load_workbook(file.file, read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Invalid Excel file.")

    ws = wb.active
    if ws is None:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "No active sheet.")

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Excel must have header + data rows.")

    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    required = {"teacher_email", "subject_code", "day_of_week", "start_time", "end_time"}
    missing = required - set(header)
    if missing:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, f"Missing columns: {missing}")

    col = {h: i for i, h in enumerate(header)}
    college_id = current_user["college_id"]

    created = 0
    failed_rows = []

    for row_num, row in enumerate(rows[1:], start=2):
        def cell(name):
            idx = col.get(name)
            if idx is None or idx >= len(row) or row[idx] is None:
                return None
            return str(row[idx]).strip()

        teacher_email = cell("teacher_email")
        subject_code  = cell("subject_code")
        day_str       = cell("day_of_week")
        start         = cell("start_time")
        end           = cell("end_time")
        room          = cell("room")
        section_name  = cell("section_name")

        if not all([teacher_email, subject_code, day_str, start, end]):
            failed_rows.append({"row": row_num, "reason": "Missing required data"})
            continue

        # Validate day
        try:
            day = DayOfWeek(day_str.lower())
        except ValueError:
            failed_rows.append({"row": row_num, "reason": f"Invalid day '{day_str}'"})
            continue

        # Find teacher
        teacher = db.query(User).filter(
            User.email == teacher_email,
            User.college_id == college_id,
            User.role.in_([UserRole.teacher, UserRole.hod, UserRole.principal]),
        ).first()
        if not teacher:
            failed_rows.append({"row": row_num, "reason": f"Teacher '{teacher_email}' not found"})
            continue

        # Find subject
        subj = db.query(Subject).filter(Subject.code == subject_code).first()
        if not subj:
            failed_rows.append({"row": row_num, "reason": f"Subject '{subject_code}' not found"})
            continue

        # Find section (optional)
        sec_id = None
        if section_name:
            sec = db.query(Section).filter(Section.name == section_name).first()
            if sec:
                sec_id = sec.id

        # Check teacher clash
        clash = db.query(Timetable).filter(
            Timetable.teacher_id == teacher.id,
            Timetable.day_of_week == day,
            Timetable.start_time < end,
            Timetable.end_time > start,
        ).first()
        if clash:
            failed_rows.append({"row": row_num, "reason": f"Teacher clash on {day.value} {start}–{end}"})
            continue

        # Check exact duplicate
        dup = db.query(Timetable).filter(
            Timetable.subject_id == subj.id,
            Timetable.day_of_week == day,
            Timetable.start_time == start,
        ).first()
        if dup:
            failed_rows.append({"row": row_num, "reason": "Duplicate entry"})
            continue

        db.add(Timetable(
            subject_id=subj.id,
            teacher_id=teacher.id,
            day_of_week=day,
            start_time=start,
            end_time=end,
            room=room,
            section_id=sec_id,
        ))
        created += 1

    db.commit()
    wb.close()

    logger.info("📅 TIMETABLE BULK-UPLOAD │ created=%d │ failed=%d │ by user_id=%d",
                created, len(failed_rows), current_user["id"])

    return {"created": created, "failed_rows": failed_rows}


# ══════════════════════════════════════════════════════════════════════
# GET /api/timetable/export-excel
# ══════════════════════════════════════════════════════════════════════

@router.get("/export-excel")
def export_excel(
    department_id: Optional[int] = None,
    current_user:  dict    = Depends(hod_or_above),
    db:            Session = Depends(get_db),
):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    dept_id = department_id or current_user.get("department_id")
    if not dept_id:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "department_id required.")

    dept = db.query(Department).filter(Department.id == dept_id).first()

    course_ids = [r[0] for r in db.query(Course.id).filter(Course.department_id == dept_id).all()]
    subject_ids = [r[0] for r in db.query(Subject.id).filter(Subject.course_id.in_(course_ids)).all()] if course_ids else []

    entries = (
        db.query(Timetable)
        .filter(Timetable.subject_id.in_(subject_ids))
        .order_by(Timetable.day_of_week, Timetable.start_time)
        .all()
    ) if subject_ids else []

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Timetable"

    # Header
    header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
    header_font = Font(bold=True, color="ffffff", size=11)

    headers = ["Day", "Period", "Start Time", "End Time", "Subject", "Code",
               "Teacher", "Section", "Room", "Lab"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for i, e in enumerate(entries, 2):
        subj = db.query(Subject).filter(Subject.id == e.subject_id).first()
        teacher = db.query(User).filter(User.id == e.teacher_id).first()
        sec_name = ""
        if e.section_id:
            sec = db.query(Section).filter(Section.id == e.section_id).first()
            sec_name = sec.name if sec else ""

        ws.cell(row=i, column=1, value=e.day_of_week.value.capitalize())
        ws.cell(row=i, column=2, value=e.period_number or "")
        ws.cell(row=i, column=3, value=e.start_time)
        ws.cell(row=i, column=4, value=e.end_time)
        ws.cell(row=i, column=5, value=subj.name if subj else "")
        ws.cell(row=i, column=6, value=subj.code if subj else "")
        ws.cell(row=i, column=7, value=teacher.name if teacher else "")
        ws.cell(row=i, column=8, value=sec_name)
        ws.cell(row=i, column=9, value=e.room or "")
        ws.cell(row=i, column=10, value="Yes" if e.is_lab else "")

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 30)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    wb.close()

    dept_name = dept.name if dept else "department"
    logger.info("📅 TIMETABLE EXPORT │ dept=%s │ entries=%d │ by user_id=%d",
                dept_name, len(entries), current_user["id"])

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="timetable_{dept_name}.xlsx"'},
    )


# ══════════════════════════════════════════════════════════════════════
# GET /api/timetable/my-today — teacher's schedule today
# ══════════════════════════════════════════════════════════════════════

@router.get("/my-today")
def my_today(
    current_user: dict    = Depends(teacher_or_above),
    db:           Session = Depends(get_db),
):
    today = _today_day()
    if today is None:
        return []  # Sunday

    entries = (
        db.query(Timetable)
        .filter(
            Timetable.teacher_id == current_user["id"],
            Timetable.day_of_week == today,
        )
        .order_by(Timetable.start_time)
        .all()
    )

    return [_serialize_entry(e, db, include_session=True) for e in entries]


# ══════════════════════════════════════════════════════════════════════
# GET /api/timetable/my-week — full week schedule
# ══════════════════════════════════════════════════════════════════════

@router.get("/my-week")
def my_week(
    current_user: dict    = Depends(teacher_or_above),
    db:           Session = Depends(get_db),
):
    entries = (
        db.query(Timetable)
        .filter(Timetable.teacher_id == current_user["id"])
        .order_by(Timetable.day_of_week, Timetable.start_time)
        .all()
    )

    by_day = {d: [] for d in DAY_ORDER}
    for e in entries:
        key = e.day_of_week.value
        if key in by_day:
            by_day[key].append(_serialize_entry(e, db, include_session=True))

    return {
        d: by_day[d]
        for d in DAY_ORDER
    }


# ══════════════════════════════════════════════════════════════════════
# GET /api/timetable/my-current-class — class happening right now
# ══════════════════════════════════════════════════════════════════════

@router.get("/my-current-class")
def my_current_class(
    current_user: dict    = Depends(teacher_or_above),
    db:           Session = Depends(get_db),
):
    today = _today_day()
    if today is None:
        return None  # Sunday

    now_str = datetime.now(tz=timezone.utc).strftime("%H:%M")

    entry = (
        db.query(Timetable)
        .filter(
            Timetable.teacher_id == current_user["id"],
            Timetable.day_of_week == today,
            Timetable.start_time <= now_str,
            Timetable.end_time > now_str,
        )
        .first()
    )

    if not entry:
        return None

    result = _serialize_entry(entry, db, include_session=True)

    # can_start: no active session yet
    result["can_start_attendance"] = result["session_status"] == "not_started"
    # can_end: session is active
    result["can_end_attendance"] = result["session_status"] == "active"

    return result


# ══════════════════════════════════════════════════════════════════════
# POST /api/timetable/start-from-timetable/{timetable_id}
# ══════════════════════════════════════════════════════════════════════

@router.post("/start-from-timetable/{timetable_id}")
def start_from_timetable(
    timetable_id: int,
    body:         StartFromTimetableRequest,
    current_user: dict    = Depends(teacher_or_above),
    db:           Session = Depends(get_db),
):
    entry = db.query(Timetable).filter(Timetable.id == timetable_id).first()
    if not entry:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Timetable entry not found.")

    if entry.teacher_id != current_user["id"]:
        raise HTTPException(status.HTTP_403_FORBIDDEN,
                            "This timetable entry does not belong to you.")

    subject = db.query(Subject).filter(Subject.id == entry.subject_id).first()
    if not subject:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Subject not found.")

    today = date.today()

    # Check for existing active session
    dup_filters = [
        AttendanceSession.subject_id == entry.subject_id,
        AttendanceSession.date == today,
        AttendanceSession.status == SessionStatus.active,
    ]
    if entry.section_id:
        dup_filters.append(AttendanceSession.section_id == entry.section_id)
    else:
        dup_filters.append(AttendanceSession.section_id.is_(None))

    existing = db.query(AttendanceSession).filter(*dup_filters).first()
    if existing:
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            f"An active session already exists for this subject today (id={existing.id}).",
        )

    # Generate secrets
    qr_secret = secrets.token_hex(32)
    bt_token = generate_bluetooth_token()
    now = datetime.now(tz=timezone.utc)

    session = AttendanceSession(
        subject_id=entry.subject_id,
        teacher_id=current_user["id"],
        section_id=entry.section_id,
        date=today,
        start_time=now.time(),
        status=SessionStatus.active,
        teacher_latitude=body.teacher_latitude,
        teacher_longitude=body.teacher_longitude,
        bluetooth_token=bt_token,
        qr_secret=qr_secret,
    )
    db.add(session)
    db.flush()

    # Enroll students
    student_filters = [
        User.course_id == subject.course_id,
        User.semester == subject.semester,
        User.role == UserRole.student,
        User.is_active == True,
    ]
    if entry.section_id:
        student_filters.append(User.section_id == entry.section_id)

    students = db.query(User).filter(*student_filters).all()

    for student in students:
        db.add(AttendanceRecord(
            session_id=session.id,
            student_id=student.id,
            status=AttendanceStatus.absent,
            marked_by=MarkedBy.auto_absent,
            face_verified=False,
            gps_verified=False,
            bluetooth_verified=False,
        ))

    session.total_students = len(students)
    db.commit()

    logger.info("📅 TIMETABLE START │ timetable_id=%d │ session_id=%d │ subj=%s │ %d students │ teacher_id=%d",
                timetable_id, session.id, subject.name, len(students), current_user["id"])

    # Push notifications
    teacher_user = db.query(User).filter(User.id == current_user["id"]).first()
    t_name = teacher_user.name if teacher_user else "Your teacher"
    student_ids = [s.id for s in students]
    if student_ids:
        send_push_to_many(
            user_ids=student_ids,
            title=f"📢 {subject.name} — Attendance Started",
            body=f"{t_name} has started attendance for {subject.name}. Scan QR now!",
            db=db,
            data={"type": "session_started", "session_id": session.id, "screen": "ScanQR"},
        )

    return {
        "session_id":      session.id,
        "subject_name":    subject.name,
        "subject_code":    subject.code,
        "bluetooth_token":  bt_token,
        "qr_secret_hint":  qr_secret[:8],
        "total_students":  len(students),
        "started_at":      now.isoformat(),
        "timetable_id":    timetable_id,
    }


# ══════════════════════════════════════════════════════════════════════
# GET /api/timetable/my-section-timetable — student's section timetable
# ══════════════════════════════════════════════════════════════════════

@router.get("/my-section-timetable")
def my_section_timetable(
    current_user: dict    = Depends(any_authenticated),
    db:           Session = Depends(get_db),
):
    user = db.query(User).filter(User.id == current_user["id"]).first()
    if not user or not user.section_id:
        return {"timetable": []}

    entries = (
        db.query(Timetable)
        .filter(Timetable.section_id == user.section_id)
        .order_by(Timetable.day_of_week, Timetable.start_time)
        .all()
    )

    by_day = {d: [] for d in DAY_ORDER}
    for e in entries:
        key = e.day_of_week.value
        if key in by_day:
            by_day[key].append(_serialize_entry(e, db))

    return {
        "timetable": [
            {"day": DAY_NAMES[d], "entries": by_day[d]}
            for d in DAY_ORDER if by_day[d]
        ]
    }
